from openpyxl import styles, load_workbook
from datetime import datetime, timedelta
import pandas as pd
import directorypaths as dp
from openpyxl.styles import Alignment, Font
from openpyxl.styles.borders import BORDER_THIN
from openpyxl.styles import Border, Side


input_date_format = '%d%m%y'
month_name = {"01": "JAN", "02": "FEB", "03": "MAR", "04": "APR", "05": "MAY"}

header1 = {'SYMBOL': 'Scrip Details', 'NAME': ' ', 'HI_52_WK_x': '52 Week', 'LO_52_WK_x': ' ',
           'LAST_MO_HIGH_PRICE': 'Last Month', 'LAST_MO_LOW_PRICE': ' ', 'MO_HIGH_PRICE': 'Current Month',
           'MO_LOW_PRICE': ' ', 'LAST_WEEK_HIGH_PRICE': 'Last Week', 'LAST_WEEK_LOW_PRICE': ' ',
           'WE_HIGH_PRICE': 'Current Week', 'WE_LOW_PRICE': ' '}

existing_header = {'SYMBOL': 'Scrip Details', 'High': '52 Week ', 'High.1': 'Last Month',
                   'High.2': 'Current Month', 'High.3': 'Last Week', 'High.4': 'Current Week'}

dynamic_header = {'OPEN_PRICE': 'OPEN_PRICE', 'HIGH_PRICE': ' ', 'LOW_PRICE': ' ', 'CLOSE_PRICE': ' ',
                  'NET_TRDQTY': ' '}

header3 = ['SYMBOL', 'Name', 'High', 'Low', 'High', 'Low', 'High', 'Low', 'High', 'Low', 'High', 'Low', 'Open', 'High',
           'Low', 'Close', 'Change (%)', 'Volume']

date_wise_dynamic_header = ['Open', 'High', 'Low', 'Close', 'Change (%)', 'Volume']

font_header = Font(name='Arial',
                   size=8,
                   bold=True,
                   italic=False,
                   vertAlign=None,
                   underline='none',
                   strike=False,
                   color='FF000000')

align_header = Alignment(horizontal='center',
                         vertical='center',
                         text_rotation=0,
                         wrap_text=False,
                         shrink_to_fit=True,
                         indent=0)

font_body = Font(name='Arial',
                 size=8,
                 bold=False,
                 italic=False,
                 vertAlign=None,
                 underline='none',
                 strike=False,
                 color='FF000000')

align_body_str = Alignment(horizontal='left',
                           vertical='center',
                           text_rotation=0,
                           wrap_text=False,
                           shrink_to_fit=False,
                           indent=0)

align_body_num = Alignment(horizontal='right',
                           vertical='center',
                           text_rotation=0,
                           wrap_text=False,
                           shrink_to_fit=False,
                           indent=0)


thin_border = Border(
    left=Side(border_style=BORDER_THIN, color='00000000'),
    right=Side(border_style=BORDER_THIN, color='00000000'),
    top=Side(border_style=BORDER_THIN, color='00000000'),
    bottom=Side(border_style=BORDER_THIN, color='00000000')
)


red_color = 'F6646B'
green_color = 'CAFF33'
red_fill = styles.PatternFill(start_color=red_color, end_color=red_color, fill_type='solid')
green_fill = styles.PatternFill(start_color=green_color, end_color=green_color, fill_type='solid')


# get the current date from CurrentDate.txt file
def get_current_date():
    current_date_file = open("CurrentDate.txt")
    current_date = current_date_file.read()
    current_date_file.close()
    return current_date


# Create excel sheet name from given Date.
# If the date is 160419, the Sheet name will be APR2019
# If the date is 070519, the sheet name will be MAY2019
def create_sheet_name(current_date_str):
    return month_name[current_date_str[2:4]] + '20' + current_date_str[4:6]


# Create the date from given String in specified input_date_format
# and return the date. Example, if the date format is %d%m%y and
# given date is 100419, then the date created and returned is 2019-04-10
def create_date(current_date):
    return datetime.strptime(current_date, input_date_format)


# This method checks if a sheet exists in an excel
# Returns True if the sheet exist,
# False,if the File or Sheet does not exist.
def check_sheet_exist(excel_path, sheet_name):
    """
    Check if a sheet exist in the given excel.
    :param excel_path: The path of the excel file
    :param sheet_name: The sheet name
    :return: True if the sheet exists. False if the File or Sheet does not exist.
    """
    try:
        wb = load_workbook(excel_path)
        sheet = wb[sheet_name]
    except KeyError:
        return False
    except FileNotFoundError:
        return False

    return True


# Append the new data in existing records
def append_df_to_excel(filename, df, sheet_name='Sheet1', startrow=None,
                       truncate_sheet=False,
                       **to_excel_kwargs):
    """
    Append a DataFrame [df] to existing Excel file [filename]
    into [sheet_name] Sheet.
    If [filename] doesn't exist, then this function will create it.

    Parameters:
      filename : File path or existing ExcelWriter
                 (Example: '/path/to/file.xlsx')
      df : dataframe to save to workbook
      sheet_name : Name of sheet which will contain DataFrame.
                   (default: 'Sheet1')
      startrow : upper left cell row to dump data frame.
                 Per default (startrow=None) calculate the last row
                 in the existing DF and write to the next row...
      truncate_sheet : truncate (remove and recreate) [sheet_name]
                       before writing DataFrame to Excel file
      to_excel_kwargs : arguments which will be passed to `DataFrame.to_excel()`
                        [can be dictionary]

    Returns: None
    """

    # ignore [engine] parameter if it was passed
    if 'engine' in to_excel_kwargs:
        to_excel_kwargs.pop('engine')

    writer = pd.ExcelWriter(filename, engine='openpyxl')

    # Python 2.x: define [FileNotFoundError] exception if it doesn't exist
    try:
        FileNotFoundError
    except NameError:
        FileNotFoundError = IOError


    try:
        # try to open an existing workbook
        writer.book = load_workbook(filename)

        # get the last row in the existing Excel sheet
        # if it was not specified explicitly
        if startrow is None and sheet_name in writer.book.sheetnames:
            startrow = writer.book[sheet_name].max_row

        # truncate sheet
        if truncate_sheet and sheet_name in writer.book.sheetnames:
            # index of [sheet_name] sheet
            idx = writer.book.sheetnames.index(sheet_name)
            # remove [sheet_name]
            writer.book.remove(writer.book.worksheets[idx])
            # create an empty sheet [sheet_name] using old index
            writer.book.create_sheet(sheet_name, idx)

        # copy existing sheets
        writer.sheets = {ws.title:ws for ws in writer.book.worksheets}
    except FileNotFoundError:
        # file does not exist yet, we will create it
        pass

    if startrow is None:
        startrow = 0

    # write out the new sheet
    if startrow == 0:
        df.to_excel(writer, sheet_name, startrow=startrow, index=False, **to_excel_kwargs)
    else:
        df.to_excel(writer, sheet_name, startrow=startrow, header=None, index=False, **to_excel_kwargs)
    # save the workbook
    writer.save()


current_date_str = get_current_date()
gl_sheet_name = create_sheet_name(current_date_str)
gl_master_data_sheet_name = 'details'
gl_formatted_date = create_date(current_date_str)

red_color = 'F6646B'
green_color = 'CAFF33'
red_fill = styles.PatternFill(start_color=red_color, end_color=red_color, fill_type='solid')
green_fill = styles.PatternFill(start_color=green_color, end_color=green_color, fill_type='solid')


# Check if a given date is holiday or not
# If the date is weekend, then it is holiday
# Else, if the date is there in the given holiday list
# the Holiday List will be given in NSE_Holiday_List.xlsx file.
def check_holiday(current_date):
    if current_date.weekday() == 5 or current_date.weekday() == 6:
        return True

    return current_date in pd.read_excel(dp.nse_holiday_list)['Date'].tolist()


# Check whether a given date is first day of week or not
def check_new_week(current_date):
    if current_date.weekday() == 0:
        return True

    if check_holiday(current_date):
        return False

    no_of_day = current_date.weekday()

    while no_of_day != 0:
        current_date = current_date - timedelta(days=1)
        if check_holiday(current_date):
            no_of_day = current_date.weekday()
            continue
        else:
            return False

    return True


# Returns the last business day offset by given no of days
def offset_business_day(current_date, offset_days):

    curr_offset_days = 0

    while curr_offset_days != offset_days:
        previous_date = current_date - timedelta(days=1)
        while check_holiday(previous_date):
            previous_date = previous_date - timedelta(days=1)
        current_date = previous_date
        curr_offset_days = curr_offset_days + 1
    return previous_date


price_volume_header = ['SYMBOL', 'Name', 'Previous Close', 'Close Price', 'Change', 'Change(%)','Volume Change',
                       'Volume Change(%)', 'Volume', 'Prev. Volume']

week_high_low_haeder = ['SYMBOL', 'Name', '52 Week High', '52 Week Low', 'Previous Close Price',
                        'Open Price', 'High Price', 'Low Price', 'Close Price']

volatility_header = ['SYMBOL', 'Name', 'Previous Close Price', 'Open Price', 'High Price', 'Low Price', 'Close Price',
                     'Volatility', 'Volatility(%)']

volatility_header_updated = {'SYMBOL': 'SYMBOL', 'NAME': 'Name', 'PREV_CL_PR': 'Previous Close Price',
                             'OPEN_PRICE': 'Open Price', 'HIGH_PRICE': 'High Price', 'LOW_PRICE': 'Low Price',
                             'CLOSE_PRICE': 'Close Price'}

cons_increased_header1 = {'PREV_CL_PR': 'Previous Closing Price', 'CLOSE_PRICE_x': 'Closing Price',
                          'NET_TRDQTY_x': 'Volume', 'CLOSE_PRICE_y': 'Closing Price', 'NET_TRDQTY_y': 'Volume',
                          'CLOSE_PRICE': 'Closing Price', 'NET_TRDQTY': 'Volume'}

cons_increased_header2 = {' ': 'SYMBOL', 'TRADE_DATE': 'Name'}
