import pandas as pd
from pandas.errors import EmptyDataError

import stockreportgeneratorhelper as srgh
from openpyxl.styles import PatternFill
from openpyxl import load_workbook
from datetime import timedelta


class DailyReportGenerator:

    def __init__(self, input_file_name, data_sheet_name, current_date_str, config):
        self.input_file_name = input_file_name
        self.data_sheet_name = data_sheet_name
        self.current_date_str = current_date_str
        self.config = config

    # Generate Daily Reports
    # Generate the Price Volume Report
    def generate_daily_reports(self):
        current_date = srgh.create_date(self.current_date_str)
        previous_date = srgh.offset_business_day(current_date, 1, self.config)
        report_name = self.config.output_file_path + self.config.daily_report_name + '_' + str(current_date.date()) + '.xlsx'

        DailyReportGenerator.generate_price_volume_report = staticmethod(
                                   DailyReportGenerator.generate_price_volume_report)
        DailyReportGenerator.generate_price_volume_report(self, current_date, previous_date, report_name)

        DailyReportGenerator.generate_new_52_week_high_low_report = staticmethod(
                               DailyReportGenerator.generate_new_52_week_high_low_report)
        DailyReportGenerator.generate_new_52_week_high_low_report(self, current_date, report_name)

        DailyReportGenerator.generate_volatile_stock_day = staticmethod(DailyReportGenerator.generate_volatile_stock_day)
        DailyReportGenerator.generate_volatile_stock_day(self, current_date, report_name)

        DailyReportGenerator.generate_trending_scrip_list = staticmethod(DailyReportGenerator.generate_trending_scrip_list)
        DailyReportGenerator.generate_trending_scrip_list(self, current_date, report_name)

        DailyReportGenerator.generate_trending_scrip_list_2 = staticmethod(DailyReportGenerator.generate_trending_scrip_list_2)
        DailyReportGenerator.generate_trending_scrip_list_2(self, current_date, report_name)

        DailyReportGenerator.generate_ca_records = staticmethod(DailyReportGenerator.generate_ca_records)
        DailyReportGenerator.generate_ca_records(self, report_name)

        DailyReportGenerator.generate_bulk_deal_records = staticmethod(DailyReportGenerator.generate_bulk_deal_records)
        DailyReportGenerator.generate_bulk_deal_records(self, report_name)

        DailyReportGenerator.generate_block_deal_records = staticmethod(DailyReportGenerator.generate_block_deal_records)
        DailyReportGenerator.generate_block_deal_records(self, report_name)

    # Generate the report for the shares whose close price is increased
    # or decreased three consecutive days.
    def generate_trending_scrip_list(self, current_date, report_name):
        master_data = pd.read_excel(self.input_file_name, self.data_sheet_name)
        DailyReportGenerator.generate_three_cons_report(self, current_date, report_name)
        DailyReportGenerator.generate_seven_cons_report(self, master_data, current_date, report_name)

    # Three consecutive days
    def generate_three_cons_report(self, current_date, report_name):
        master_data = pd.read_excel(self.input_file_name, self.data_sheet_name)

        to_date = srgh.create_date(self.current_date_str)
        from_date = srgh.offset_business_day(current_date, 2, self.config)
        DailyReportGenerator.generate_consecutive_records(self, master_data, from_date, to_date, report_name, 3, True,
                                                          'Scrips with Price Increased Three Consecutive Session')
        DailyReportGenerator.generate_consecutive_records(self, master_data, from_date, to_date, report_name, 3, False,
                                                          'Scrips with Price Decreased Three Consecutive Session')

    # Seven consecutive days
    def generate_seven_cons_report(self, master_data, current_date, report_name):
        to_date = srgh.create_date(self.current_date_str)
        from_date = srgh.offset_business_day(current_date, 6, self.config)
        DailyReportGenerator.generate_consecutive_records(self, master_data, from_date, to_date, report_name, 7, True,
                                                          'Scrips with Price Increased Seven Consecutive Session')
        DailyReportGenerator.generate_consecutive_records(self, master_data, from_date, to_date, report_name, 7, False,
                                                          'Scrips with Price Decreased Seven Consecutive Session')

    # fetch_cons_records
    def generate_consecutive_records(self, master_data, from_date, to_date, report_name, frequency, is_increased,
                                     header_msg):

        temp_df = DailyReportGenerator.generate_consecutive_increased_record(self, master_data, from_date, to_date,
                                                                             report_name, frequency, is_increased)
        # check of Trendies Technical - I sheet exists
        if not temp_df.empty:
            sheet_exists = srgh.check_sheet_exist(report_name, "Trendies Technical - I")
            if not sheet_exists:
                DailyReportGenerator.create_trending_technical_1(self, report_name, "Trendies Technical - I", temp_df,
                                                                 header_msg)
            else:
                DailyReportGenerator.update_trending_technical_1(self, report_name, "Trendies Technical - I", from_date,
                                                                 temp_df, header_msg, frequency, is_increased)

    # Generate Consecutive Increased Records
    def generate_consecutive_increased_record(self, master_data, from_date, to_date, report_name, frequency,
                                              is_increased):

        if is_increased:
            price_data = master_data[(master_data['TRADE_DATE'] <= to_date) &
                                           (master_data['TRADE_DATE'] >= from_date) &
                                           (master_data['CLOSE_PRICE'] >= master_data['PREV_CL_PR'])]
        else:
            price_data = master_data[(master_data['TRADE_DATE'] <= to_date) &
                                               (master_data['TRADE_DATE'] >= from_date) &
                                               (master_data['CLOSE_PRICE'] <= master_data['PREV_CL_PR'])]
        price_data['freq'] = price_data.groupby('SYMBOL')['SYMBOL'].transform('count').copy(deep=True)
        price_data = price_data[(price_data['freq'] == frequency)].copy(deep=True)
        price_data = price_data[['SYMBOL', 'NAME', 'TRADE_DATE', 'PREV_CL_PR', 'CLOSE_PRICE', 'NET_TRDQTY']].copy(deep=True)
        temp_df = price_data[(price_data['TRADE_DATE'] == from_date)]

        if not temp_df.empty:
            app_date = from_date
            while app_date < to_date:
                app_date = app_date + timedelta(days=1)
                while srgh.check_holiday(app_date, self.config):
                    app_date = app_date + timedelta(days=1)

                temp_df1 = price_data[(price_data['TRADE_DATE'] == app_date)]
                temp_df1 = temp_df1[['SYMBOL', 'NAME', 'CLOSE_PRICE', 'NET_TRDQTY']]
                try:
                    temp_df = pd.merge(temp_df, temp_df1, left_on=['SYMBOL', 'NAME'], right_on=['SYMBOL', 'NAME'])
                    if frequency == 3:
                        temp_df.rename(columns=srgh.cons_increased_header3, inplace=True)
                    if frequency == 7:
                        # if temp_df.empty:
                        #     temp_df.rename(columns=srgh.trendies_1_empty_dataframe_header, inplace=True)
                        # else:
                        temp_df.rename(columns=srgh.cons_increased_header7, inplace=True)

                except IndexError:
                    temp_df = temp_df if not temp_df.empty else temp_df1

                while srgh.check_holiday(app_date, self.config):
                    app_date = app_date + timedelta(days=1)
            temp_df.drop(columns=['TRADE_DATE'], inplace=True)
        return temp_df

    # Create trending technical 1 sheet
    def create_trending_technical_1(self, report_name, sheet_name, temp_df, header_msg):
        book = load_workbook(report_name)
        writer = pd.ExcelWriter(report_name, engine='openpyxl')
        writer.book = book
        writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
        temp_df.to_excel(writer, "Trendies Technical - I", startrow=2, index=False)
        writer.save()
        DailyReportGenerator.format_first_set_record = staticmethod(
            DailyReportGenerator.format_first_set_record)
        DailyReportGenerator.format_first_set_record(self, report_name, header_msg)

    # Update trending technical 1 sheet
    def update_trending_technical_1(self, report_name, sheet_name, from_date, temp_dr_df, header_msg, frequency,
                                    is_increased):
        book = load_workbook(report_name)
        sheet = book['Trendies Technical - I']

        DailyReportGenerator.update_decr_scrip_list = staticmethod(
            DailyReportGenerator.update_decr_scrip_list)
        DailyReportGenerator.update_decr_scrip_list(self, temp_dr_df, sheet.max_row + 3, report_name)

        DailyReportGenerator.format_further_record_set = staticmethod(
            DailyReportGenerator.format_further_record_set)
        DailyReportGenerator.format_further_record_set(self, sheet.max_row + 2, from_date, report_name, header_msg,
                                                       frequency, is_increased)

    # Update
    def update_decr_scrip_list(self, decreased_scrip_df, star_row, report_name):
        book = load_workbook(report_name)
        writer = pd.ExcelWriter(report_name, engine='openpyxl')
        writer.book = book
        writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
        # selected_list.to_excel(writer, sheet_name, header=None, index=False)

        decreased_scrip_df.to_excel(writer, 'Trendies Technical - I',startrow=star_row, index=False)
        writer.save()

    # Format Cons Increase Report
    def format_first_set_record(self, report_name, header_msg):
        to_date = srgh.create_date(self.current_date_str)
        from_date = srgh.offset_business_day(to_date, 2, self.config)

        book = load_workbook(report_name)
        sheet = book['Trendies Technical - I']
        sheet.merge_cells('A1:I1')
        sheet.cell(row=1, column=1).value = header_msg
        sheet.merge_cells('A2:B2')
        sheet.cell(row=2, column=1).value = 'Scrip Details'
        max_column = sheet.max_column
        curr_column = 4
        while curr_column < max_column:
            sheet.merge_cells(start_row=2, start_column=curr_column, end_row=2, end_column=curr_column + 1)

            while srgh.check_holiday(from_date, self.config):
                from_date = from_date + timedelta(days=1)

            sheet.cell(row=2, column=curr_column).value = from_date.date()
            curr_column = curr_column + 2
            from_date = from_date + timedelta(days=1)

        curr_column = 1
        max_rows = sheet.max_row
        curr_row = 1
        while curr_row <= max_rows:
            while curr_column <= max_column:
                sheet.cell(curr_row, curr_column).border = srgh.thin_border
                if curr_row == 3 or curr_row == 2:
                    sheet.cell(curr_row, curr_column).font = srgh.font_header
                    sheet.cell(curr_row, curr_column).fill = PatternFill(start_color='D3D3D3', fill_type="solid")
                    sheet.cell(curr_row, curr_column).alignment = srgh.align_header
                elif curr_row == 1:
                    sheet.cell(curr_row, curr_column).font = srgh.font_header
                    sheet.cell(curr_row, curr_column).fill = PatternFill(start_color='CAFF33', fill_type="solid")
                    sheet.cell(curr_row, curr_column).alignment = srgh.align_header
                else:
                    sheet.cell(curr_row, curr_column).font = srgh.font_body
                    if curr_column < 3:
                        sheet.cell(curr_row, curr_column).alignment = srgh.align_body_str
                    else:
                        sheet.cell(curr_row, curr_column).alignment = srgh.align_body_num
                curr_column = curr_column + 1
            sheet.row_dimensions[curr_row].height = 20  # In pixels
            curr_row = curr_row + 1
            curr_column = 1

        curr_row_no = 4
        for rows in sheet.iter_rows(min_row=4, max_row=max_rows, min_col=1):
            for cell in rows:
                if curr_row_no % 2 == 1:
                    cell.fill = PatternFill(start_color="f7ec8f", fill_type="solid")
                else:
                    cell.fill = PatternFill(start_color="edead7", fill_type="solid")
            curr_row_no = curr_row_no + 1
        sheet.column_dimensions['A'].width = 18
        sheet.column_dimensions['B'].width = 25
        sheet.column_dimensions['C'].width = 22
        sheet.column_dimensions['D'].width = 15
        sheet.column_dimensions['E'].width = 15
        sheet.column_dimensions['F'].width = 15
        sheet.column_dimensions['G'].width = 15
        sheet.column_dimensions['H'].width = 15
        sheet.column_dimensions['I'].width = 15
        sheet.column_dimensions['J'].width = 15
        sheet.column_dimensions['K'].width = 15
        sheet.column_dimensions['L'].width = 15
        sheet.column_dimensions['M'].width = 15
        sheet.column_dimensions['N'].width = 15
        sheet.column_dimensions['O'].width = 15
        sheet.column_dimensions['P'].width = 15
        sheet.column_dimensions['Q'].width = 15

        book.save(report_name)

    # Format Cons Increase Report
    def format_further_record_set(self, start_row, from_date, report_name, header_msg, frequency, is_increased):
        to_date = srgh.create_date(self.current_date_str)

        book = load_workbook(report_name)

        sheet = book['Trendies Technical - I']
        column_name = DailyReportGenerator.get_column_name(self, frequency)
        cell_range = 'A' + str(start_row) + ':' + column_name + str(start_row)
        sheet.merge_cells(cell_range)
        sheet.cell(row=start_row, column=1).value = header_msg
        cell_range = 'A' + str(start_row + 1) + ':'+'B'+str(start_row + 1)
        sheet.merge_cells(cell_range)
        sheet.cell(row=start_row + 1, column=1).value = 'Scrip Details'
        # max_column = sheet.max_column
        max_column = (frequency * 2) + 3
        curr_column = 4
        while curr_column < max_column:
            sheet.merge_cells(start_row=start_row + 1, start_column=curr_column, end_row=start_row + 1,
                              end_column=curr_column + 1)

            while srgh.check_holiday(from_date, self.config):
                from_date = from_date + timedelta(days=1)

            sheet.cell(row=start_row + 1, column=curr_column).value = from_date.date()
            curr_column = curr_column + 2
            from_date = from_date + timedelta(days=1)

        curr_column = 1
        max_rows = sheet.max_row
        curr_row = start_row
        while curr_row <= max_rows:
            while curr_column <= max_column:
                sheet.cell(curr_row, curr_column).border = srgh.thin_border
                if curr_row == start_row + 1 or curr_row == start_row + 2:
                    sheet.cell(curr_row, curr_column).font = srgh.font_header
                    sheet.cell(curr_row, curr_column).fill = PatternFill(start_color='D3D3D3', fill_type="solid")
                    sheet.cell(curr_row, curr_column).alignment = srgh.align_header
                elif curr_row == start_row:
                    sheet.cell(curr_row, curr_column).font = srgh.font_header
                    if is_increased:
                        sheet.cell(curr_row, curr_column).fill = PatternFill(start_color='CAFF33', fill_type="solid")
                    else:
                        sheet.cell(curr_row, curr_column).fill = PatternFill(start_color='F6646B', fill_type="solid")
                    sheet.cell(curr_row, curr_column).alignment = srgh.align_header
                else:
                    sheet.cell(curr_row, curr_column).font = srgh.font_body
                    if curr_column < 3:
                        sheet.cell(curr_row, curr_column).alignment = srgh.align_body_str
                    else:
                        sheet.cell(curr_row, curr_column).alignment = srgh.align_body_num
                curr_column = curr_column + 1
            sheet.row_dimensions[curr_row].height = 20  # In pixels
            curr_row = curr_row + 1
            curr_column = 1

        curr_row_no = start_row + 3
        for rows in sheet.iter_rows(min_row=start_row + 3, max_row=max_rows, min_col=1):
            for cell in rows:
                if curr_row_no % 2 == 1:
                    cell.fill = PatternFill(start_color="f7ec8f", fill_type="solid")
                else:
                    cell.fill = PatternFill(start_color="edead7", fill_type="solid")
            curr_row_no = curr_row_no + 1
        sheet.column_dimensions['B'].width = 25
        sheet.column_dimensions['A'].width = 12

        book.save(report_name)

    # Return the column name corresponding to frequency
    def get_column_name(self, frequency):
        if frequency == 3:
            column_name = 'I'
        else:
            column_name = 'Q'
        return column_name

    # Generate the Price Volume Report
    def generate_price_volume_report(self, current_date, previous_date, report_name):
        master_data = pd.read_excel(self.input_file_name, self.data_sheet_name)
        master_data = master_data[(master_data['TRADE_DATE'] <= current_date) &
                                  (master_data['TRADE_DATE'] >= previous_date)]

        master_data = master_data[['SYMBOL', 'NAME', 'TRADE_DATE', 'PREV_CL_PR', 'CLOSE_PRICE', 'NET_TRDQTY']]

        # Copy the previous day's volume in current days data
        master_data['PREV_VOL'] = master_data.groupby(['SYMBOL'])['NET_TRDQTY'].shift(1)

        # Filter the current date records
        master_data = master_data[(master_data['TRADE_DATE'] == current_date)]

        change = (master_data[master_data.columns[4]] - master_data[master_data.columns[3]])
        change_per = ((change * 100) / master_data[master_data.columns[3]])

        vol_change = (master_data[master_data.columns[5]] - master_data[master_data.columns[6]])
        vol_change_per = ((vol_change * 100) / master_data[master_data.columns[6]])

        master_data.insert(5, "Change", change)
        master_data.insert(6, "Change(%)", change_per)

        master_data.insert(7, "Volume Change", vol_change)
        master_data.insert(8, "Volume Change(%)", vol_change_per)

        master_data.drop(columns=['TRADE_DATE'], inplace=True)

        master_data.sort_values(["Change(%)"], axis=0, ascending=False, inplace=True)

        price_incr_vol_incr = master_data[(master_data['Change'] >= 0) &
                                          (master_data['Volume Change'] >= 0)]

        price_incr_vol_decr = master_data[(master_data['Change'] >= 0) &
                                          (master_data['Volume Change'] <= 0)]

        price_decr_vol_incr = master_data[(master_data['Change'] <= 0) &
                                          (master_data['Volume Change'] >= 0)]

        price_decr_vol_decr = master_data[(master_data['Change'] <= 0) &
                                          (master_data['Volume Change'] <= 0)]

        narration1 = pd.DataFrame({'SYMBOL': 'Price Increased - Volume Increased', 'NAME': ' ',
                                   'PREV_CL_PR': '', 'CLOSE_PRICE': '', 'Change': '', 'Change(%)': '',
                                   'Volume Change': '', 'Volume Change(%)': '', 'NET_TRDQTY': '', 'PREV_VOL': ''},
                                  index=[0])

        cust_header = pd.DataFrame({'SYMBOL': 'SYMBOL', 'NAME': 'Name', 'PREV_CL_PR': 'Previous Close',
                                    'CLOSE_PRICE': 'Close Price', 'Change': 'Change', 'Change(%)': 'Change(%)',
                                    'Volume Change': 'Volume Change', 'Volume Change(%)': 'Volume Change(%)',
                                    'NET_TRDQTY': 'Volume', 'PREV_VOL': 'Prev. Volume'}, index=[0])

        narration2 = pd.DataFrame({'SYMBOL': 'Price Increased - Volume Decreased', 'NAME': ' ',
                                   'PREV_CL_PR': '', 'CLOSE_PRICE': '', 'Change': '', 'Change(%)': '',
                                   'Volume Change': '', 'Volume Change(%)': '', 'NET_TRDQTY': '', 'PREV_VOL': ''},
                                  index=[0])

        narration3 = pd.DataFrame({'SYMBOL': 'Price Decreased - Volume Increased', 'NAME': ' ',
                                   'PREV_CL_PR': '', 'CLOSE_PRICE': '', 'Change': '', 'Change(%)': '',
                                   'Volume Change': '', 'Volume Change(%)': '', 'NET_TRDQTY': '', 'PREV_VOL': ''},
                                  index=[0])

        narration4 = pd.DataFrame({'SYMBOL': 'Price Decreased - Volume Decreased', 'NAME': ' ',
                                   'PREV_CL_PR': '', 'CLOSE_PRICE': '', 'Change': '', 'Change(%)': '',
                                   'Volume Change': '', 'Volume Change(%)': '', 'NET_TRDQTY': '', 'PREV_VOL': ''},
                                  index=[0])

        # Append Data Frame when Price Increased and Volume Increased
        narration1 = pd.concat([narration1, cust_header])
        price_incr_vol_incr = pd.concat([narration1, price_incr_vol_incr], sort=False)

        # Append Data Frame when Price Increased and Volume Increased
        narration2 = pd.concat([narration2, cust_header])

        price_incr_vol_decr = pd.concat([narration2, price_incr_vol_decr], sort=False)
        price_incr_vol_incr = pd.concat([price_incr_vol_incr, price_incr_vol_decr], sort=False)

        # Append Data Frame when Price Increased and Volume Increased
        narration3 = pd.concat([narration3, cust_header])

        price_decr_vol_incr = pd.concat([narration3, price_decr_vol_incr], sort=False)
        price_incr_vol_incr = pd.concat([price_incr_vol_incr, price_decr_vol_incr], sort=False)

        # Append Data Frame when Price Increased and Volume Increased
        narration4 = pd.concat([narration4, cust_header])

        price_decr_vol_decr = pd.concat([narration4, price_decr_vol_decr], sort=False)
        price_incr_vol_incr = pd.concat([price_incr_vol_incr, price_decr_vol_decr], sort=False)

        price_incr_vol_incr = price_incr_vol_incr[['SYMBOL', 'NAME', 'PREV_CL_PR', 'CLOSE_PRICE', 'Change', 'Change(%)',
                                                   'PREV_VOL', 'NET_TRDQTY', 'Volume Change', 'Volume Change(%)']]\
            .copy(deep=True)

        # Format the Excel sheet
        writer = pd.ExcelWriter(report_name, engine='openpyxl')

        # Convert the dataframe to an XlsxWriter Excel object.
        price_incr_vol_incr.to_excel(writer, sheet_name=self.config.sheet_name_price_volume, header=None, index=False)

        # Get the openpyxl workbook and worksheet objects.
        worksheet = writer.sheets[self.config.sheet_name_price_volume]

        max_rows = worksheet.max_row
        st_row = 1
        st_col = 1
        for row_cells in worksheet.iter_rows(min_row=1, max_row=max_rows):
            for cell in row_cells:
                worksheet.cell(st_row, st_col).border = srgh.thin_border
                if 'Price' in str(cell.value) and 'Volume' in str(cell.value):
                    worksheet.merge_cells(start_row=st_row, start_column=st_col, end_row=st_row, end_column=st_col + 9)
                    worksheet.cell(st_row, st_col).fill = PatternFill(start_color='D3D3D3', fill_type="solid")
                    worksheet.cell(st_row, st_col).font = srgh.font_header
                    worksheet.cell(st_row, st_col).alignment = srgh.align_header
                    break
                elif str(cell.value) in srgh.price_volume_header:
                    worksheet.cell(st_row, st_col).fill = PatternFill(start_color='D3D3D3', fill_type="solid")
                    worksheet.cell(st_row, st_col).font = srgh.font_header
                    worksheet.cell(st_row, st_col).alignment = srgh.align_header
                    st_col = st_col + 1

                else:
                    worksheet.cell(st_row, st_col).font = srgh.font_body
                    if st_row % 2 == 1:
                        cell.fill = PatternFill(start_color="f7ec8f", fill_type="solid")
                    else:
                        cell.fill = PatternFill(start_color="edead7", fill_type="solid")

                    if type(cell.value) == str:
                        worksheet.cell(st_row, st_col).alignment = srgh.align_body_str
                    else:
                        worksheet.cell(st_row, st_col).alignment = srgh.align_body_num
                    st_col = st_col + 1

            st_col = 1
            st_row = st_row + 1
            worksheet.row_dimensions[st_row].height = 20  # In pixels

            worksheet.column_dimensions['A'].width = 15
            worksheet.column_dimensions['B'].width = 25
            worksheet.column_dimensions['C'].width = 15
            worksheet.column_dimensions['D'].width = 15
            worksheet.column_dimensions['E'].width = 15
            worksheet.column_dimensions['F'].width = 15
            worksheet.column_dimensions['G'].width = 15
            worksheet.column_dimensions['H'].width = 15
            worksheet.column_dimensions['I'].width = 15
            worksheet.column_dimensions['J'].width = 18
        writer.save()

    # Generate the Price Volume Report
    def generate_new_52_week_high_low_report(self, current_date, report_name):

        master_data = pd.read_excel(self.input_file_name, self.data_sheet_name)
        master_data = master_data[(master_data['TRADE_DATE'] == current_date)]

        master_data = master_data[['SYMBOL', 'NAME', 'HI_52_WK', 'LO_52_WK', 'PREV_CL_PR', 'OPEN_PRICE',
                                   'HIGH_PRICE', 'LOW_PRICE', 'CLOSE_PRICE']]

        new_52_week_high = master_data[(master_data['HI_52_WK'] == master_data['HIGH_PRICE'])]

        new_52_week_low = master_data[(master_data['LO_52_WK'] == master_data['LOW_PRICE'])]

        close_52_week_high = master_data[master_data['CLOSE_PRICE'] >
                                         (master_data['HI_52_WK'] - ((master_data['HI_52_WK'] * 10) / 100))]

        close_52_week_low = master_data[master_data['CLOSE_PRICE'] <
                                        (master_data['LO_52_WK'] + ((master_data['LO_52_WK'] * 10) / 100))]

        narration1 = pd.DataFrame({'SYMBOL': 'New 52 Week High', 'NAME': ' ',
                                   'HI_52_WK': '', 'LO_52_WK': '', 'PREV_CL_PR': '', 'OPEN_PRICE': '',
                                   'HIGH_PRICE': '', 'LOW_PRICE': '', 'CLOSE_PRICE': ''},
                                  index=[0])

        cust_header = pd.DataFrame({'SYMBOL': 'SYMBOL', 'NAME': 'Name', 'HI_52_WK': '52 Week High',
                                    'LO_52_WK': '52 Week Low', 'PREV_CL_PR': 'Previous Close Price',
                                    'OPEN_PRICE': 'Open Price', 'HIGH_PRICE': 'High Price', 'LOW_PRICE': 'Low Price',
                                    'CLOSE_PRICE': 'Close Price'}, index=[0])

        narration2 = pd.DataFrame({'SYMBOL': 'New 52 Week Low', 'NAME': ' ',
                                   'HI_52_WK': '', 'LO_52_WK': '', 'PREV_CL_PR': '', 'OPEN_PRICE': '',
                                   'HIGH_PRICE': '', 'LOW_PRICE': '', 'CLOSE_PRICE': ''},
                                  index=[0])

        narration3 = pd.DataFrame({'SYMBOL': 'Near 52 Week High (10%)', 'NAME': ' ',
                                   'HI_52_WK': '', 'LO_52_WK': '', 'PREV_CL_PR': '', 'OPEN_PRICE': '',
                                   'HIGH_PRICE': '', 'LOW_PRICE': '', 'CLOSE_PRICE': ''},
                                  index=[0])

        narration4 = pd.DataFrame({'SYMBOL': 'Near 52 Week Low (10%)', 'NAME': ' ',
                                   'HI_52_WK': '', 'LO_52_WK': '', 'PREV_CL_PR': '', 'OPEN_PRICE': '',
                                   'HIGH_PRICE': '', 'LOW_PRICE': '', 'CLOSE_PRICE': ''},
                                  index=[0])

        # Append Data Frame when Price Increased and Volume Increased
        narration1 = pd.concat([narration1, cust_header])
        new_52_week_high = pd.concat([narration1, new_52_week_high], sort=False)

        # Append Data Frame when Price Increased and Volume Increased
        narration2 = pd.concat([narration2, cust_header])
        new_52_week_low = pd.concat([narration2, new_52_week_low], sort=False)
        new_52_week_high = pd.concat([new_52_week_high, new_52_week_low], sort=False)

        # Append Data Frame when Price Increased and Volume Increased
        narration3 = pd.concat([narration3, cust_header])
        close_52_week_high = pd.concat([narration3, close_52_week_high], sort=False)
        new_52_week_high = pd.concat([new_52_week_high, close_52_week_high], sort=False)

        # Append Data Frame when Price Increased and Volume Increased
        narration4 = pd.concat([narration4, cust_header])
        close_52_week_low = pd.concat([narration4, close_52_week_low], sort=False)
        new_52_week_high = pd.concat([new_52_week_high, close_52_week_low], sort=False)

        book = load_workbook(report_name)
        writer = pd.ExcelWriter(report_name, engine='openpyxl')
        writer.book = book
        writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
        new_52_week_high.to_excel(writer, "52-week-high-low", header=None, index=False)

        # Get the openpyxl workbook and worksheet objects.
        worksheet = writer.sheets['52-week-high-low']

        max_rows = worksheet.max_row
        st_row = 1
        st_col = 1
        for row_cells in worksheet.iter_rows(min_row=1, max_row=max_rows):
            for cell in row_cells:
                worksheet.cell(st_row, st_col).border = srgh.thin_border
                if 'New 52 Week' in str(cell.value) or 'Near 52 Week' in str(cell.value):
                    worksheet.merge_cells(start_row=st_row, start_column=st_col, end_row=st_row, end_column=st_col + 8)
                    worksheet.cell(st_row, st_col).fill = PatternFill(start_color='D3D3D3', fill_type="solid")
                    worksheet.cell(st_row, st_col).font = srgh.font_header
                    worksheet.cell(st_row, st_col).alignment = srgh.align_header
                    break
                elif str(cell.value) in srgh.week_high_low_haeder:
                    worksheet.cell(st_row, st_col).fill = PatternFill(start_color='D3D3D3', fill_type="solid")
                    worksheet.cell(st_row, st_col).font = srgh.font_header
                    worksheet.cell(st_row, st_col).alignment = srgh.align_header
                    st_col = st_col + 1

                else:
                    worksheet.cell(st_row, st_col).font = srgh.font_body
                    if st_row % 2 == 1:
                        cell.fill = PatternFill(start_color="f7ec8f", fill_type="solid")
                    else:
                        cell.fill = PatternFill(start_color="edead7", fill_type="solid")

                    if type(cell.value) == str:
                        worksheet.cell(st_row, st_col).alignment = srgh.align_body_str
                    else:
                        worksheet.cell(st_row, st_col).alignment = srgh.align_body_num
                    st_col = st_col + 1

            st_col = 1
            st_row = st_row + 1
            worksheet.row_dimensions[st_row].height = 20  # In pixels

            worksheet.column_dimensions['A'].width = 15
            worksheet.column_dimensions['B'].width = 25
            worksheet.column_dimensions['C'].width = 15
            worksheet.column_dimensions['D'].width = 15
            worksheet.column_dimensions['E'].width = 18
            worksheet.column_dimensions['F'].width = 15
            worksheet.column_dimensions['G'].width = 15
            worksheet.column_dimensions['H'].width = 15
            worksheet.column_dimensions['I'].width = 15

        writer.save()

    # Generate the Volatile Stock of the Day Report
    def generate_volatile_stock_day(self, current_date, report_name):

        master_data = pd.read_excel(self.input_file_name, self.data_sheet_name)
        master_data = master_data[(master_data['TRADE_DATE'] == current_date)]

        master_data = master_data[['SYMBOL', 'NAME', 'PREV_CL_PR', 'OPEN_PRICE',
                                   'HIGH_PRICE', 'LOW_PRICE', 'CLOSE_PRICE']]

        high_low_diff = (master_data[master_data.columns[4]] - master_data[master_data.columns[5]])
        high_low_diff_per = ((high_low_diff * 100) / master_data[master_data.columns[5]])

        master_data.insert(5, "Volatility", high_low_diff)
        master_data.insert(6, "Volatility(%)", high_low_diff_per)

        master_data.rename(columns=srgh.volatility_header_updated, inplace=True)

        master_data = master_data[['SYMBOL', 'Name', 'Previous Close Price', 'Open Price',
                                   'High Price', 'Low Price', 'Close Price', 'Volatility', 'Volatility(%)']]

        master_data.sort_values(["Volatility(%)"], axis=0, ascending=False, inplace=True)

        book = load_workbook(report_name)
        writer = pd.ExcelWriter(report_name, engine='openpyxl')
        writer.book = book
        writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
        master_data.to_excel(writer, "Volatility", index=False)

        # Get the openpyxl workbook and worksheet objects.
        worksheet = writer.sheets['Volatility']

        max_rows = worksheet.max_row
        st_row = 1
        st_col = 1
        for row_cells in worksheet.iter_rows(min_row=1, max_row=max_rows):
            for cell in row_cells:
                worksheet.cell(st_row, st_col).border = srgh.thin_border
                if str(cell.value) in srgh.volatility_header:
                    worksheet.cell(st_row, st_col).fill = PatternFill(start_color='D3D3D3', fill_type="solid")
                    worksheet.cell(st_row, st_col).font = srgh.font_header
                    worksheet.cell(st_row, st_col).alignment = srgh.align_header
                    st_col = st_col + 1

                else:
                    worksheet.cell(st_row, st_col).font = srgh.font_body
                    if st_row % 2 == 1:
                        cell.fill = PatternFill(start_color="f7ec8f", fill_type="solid")
                    else:
                        cell.fill = PatternFill(start_color="edead7", fill_type="solid")

                    if type(cell.value) == str:
                        worksheet.cell(st_row, st_col).alignment = srgh.align_body_str
                    else:
                        worksheet.cell(st_row, st_col).alignment = srgh.align_body_num
                    st_col = st_col + 1

            st_col = 1
            st_row = st_row + 1
            worksheet.row_dimensions[st_row].height = 20  # In pixels

            worksheet.column_dimensions['A'].width = 15
            worksheet.column_dimensions['B'].width = 25
            worksheet.column_dimensions['C'].width = 18
            worksheet.column_dimensions['D'].width = 15
            worksheet.column_dimensions['E'].width = 18
            worksheet.column_dimensions['F'].width = 15
            worksheet.column_dimensions['G'].width = 15
            worksheet.column_dimensions['H'].width = 15
            worksheet.column_dimensions['I'].width = 15

            c = worksheet['A2']
            worksheet.freeze_panes = c

        writer.save()

    def generate_trending_scrip_list_2(self, current_date, report_name):
        sheet_name = srgh.create_sheet_name(self.current_date_str)
        current_data = pd.read_excel(self.config.master_report_name, sheet_name, skiprows=1)
        current_data = current_data[['SYMBOL', 'Name', 'High.1', 'Low.1', 'High.2', 'Low.2', 'High.3', 'Low.3',
                                     'High.4', 'Low.4']]

        # Read the Scrip List
        scrip_list = pd.read_excel(self.config.master_data_file_name, "Details")
        scrip_list = scrip_list[['SYMBOL', 'TRADE_DATE', 'CLOSE_PRICE']]
        scrip_list = scrip_list[(scrip_list['TRADE_DATE']) == srgh.create_date(self.current_date_str)]

        # Prepare the list for Selected List
        merge_list = pd.merge(current_data, scrip_list, on=["SYMBOL", "SYMBOL"])
        DailyReportGenerator.report_closed_above_month_week_high(self, merge_list, report_name, True, True, True,
                                                                 'Scrips closed above last month and last week high '
                                                                 'price')
        DailyReportGenerator.report_closed_above_month_week_high(self, merge_list, report_name, False, True, True,
                                                                 'Scrips closed above last month high price')
        DailyReportGenerator.report_closed_above_month_week_high(self, merge_list, report_name, True, False, True,
                                                                 'Scrips closed above last week high price')
        DailyReportGenerator.report_closed_above_month_week_high(self, merge_list, report_name, True, True, False,
                                                                 'Scrips closed below last month and last week low '
                                                                 'price')
        DailyReportGenerator.report_closed_above_month_week_high(self, merge_list, report_name, False, True, False,
                                                                 'Scrips closed below last month low price')
        DailyReportGenerator.report_closed_above_month_week_high(self, merge_list, report_name, True, False, False,
                                                                 'Scrips closed below last week low price')

    def report_closed_above_month_week_high(self, merge_list, report_name, last_week, last_month, increased,
                                            header_msg):
        selected_record = DailyReportGenerator.final_data_set(self, merge_list, last_week, last_month, increased)
        selected_record = selected_record[['SYMBOL', 'Name', 'High.1', 'Low.1', 'High.2', 'Low.2', 'High.3',
                                         'Low.3', 'High.4', 'Low.4', 'CLOSE_PRICE']]

        selected_record.rename(columns=srgh.trendies_2_header, inplace=True)
        # check of Trendies Technical - 2 sheet exists
        sheet_exists = srgh.check_sheet_exist(report_name, "Trending Technical - 2")
        if not sheet_exists:
            DailyReportGenerator.create_trending_technical_2_sheet(self, report_name, "Trending Technical - 2",
                                                                   selected_record, header_msg)
        else:
            DailyReportGenerator.update_trending_technical_2_sheet(self, report_name, "Trending Technical - 2",
                                                                   selected_record, header_msg, increased)

        # DailyReportGenerator.create_trending_technical_2_sheet(self, report_name, 'Trending Technical - 2',
        #                                                       selected_record, header_msg)

    def final_data_set(self, initial_list, last_week, last_month, increased):
        if increased:
            if last_month and last_week:
                temp_list = initial_list[(initial_list['High.1'] <= initial_list['CLOSE_PRICE'])]
                temp_list = temp_list[(temp_list['High.3'] <= temp_list['CLOSE_PRICE'])]
                return temp_list
            if last_month:
                return initial_list[(initial_list['High.1'] <= initial_list['CLOSE_PRICE'])]
            if last_week:
                return initial_list[(initial_list['High.3'] <= initial_list['CLOSE_PRICE'])]
        else:
            if last_month and last_week:
                temp_list = initial_list[(initial_list['Low.1'] > initial_list['CLOSE_PRICE'])]
                temp_list = temp_list[(temp_list['Low.3'] > temp_list['CLOSE_PRICE'])]
                return temp_list
            if last_month:
                return initial_list[(initial_list['Low.1'] > initial_list['CLOSE_PRICE'])]
            if last_week:
                return initial_list[(initial_list['Low.3'] > initial_list['CLOSE_PRICE'])]

    # Create trending technical 1 sheet
    def create_trending_technical_2_sheet(self, report_name, sheet_name, temp_df, header_msg):
        book = load_workbook(report_name)
        writer = pd.ExcelWriter(report_name, engine='openpyxl')
        writer.book = book
        writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
        temp_df.to_excel(writer, sheet_name, startrow=2, index=False)
        writer.save()
        DailyReportGenerator.format_close_above_last_week = staticmethod(
            DailyReportGenerator.format_close_above_last_week)
        DailyReportGenerator.format_close_above_last_week(self, report_name, header_msg)

    # Update trending technical 1 sheet
    def update_trending_technical_2_sheet(self, report_name, sheet_name, selected_record, header_msg, is_increased):
        book = load_workbook(report_name)
        sheet = book[sheet_name]

        DailyReportGenerator.update_selected_scrip_list = staticmethod(
            DailyReportGenerator.update_selected_scrip_list)
        DailyReportGenerator.update_selected_scrip_list(self, selected_record, sheet.max_row + 3, report_name)

        DailyReportGenerator.format_selected_scrip_list = staticmethod(
            DailyReportGenerator.format_selected_scrip_list)
        DailyReportGenerator.format_selected_scrip_list(self, sheet.max_row + 2, report_name, header_msg, is_increased)

    def update_selected_scrip_list(self, selected_record, star_row, report_name):
        book = load_workbook(report_name)
        writer = pd.ExcelWriter(report_name, engine='openpyxl')
        writer.book = book
        writer.sheets = dict((ws.title, ws) for ws in book.worksheets)

        selected_record.to_excel(writer, 'Trending Technical - 2', startrow=star_row, index=False)
        writer.save()

    # Format Cons Increase Report
    def format_selected_scrip_list(self, start_row, report_name, header_msg, is_increased):

        book = load_workbook(report_name)

        sheet = book['Trending Technical - 2']
        cell_range = 'A' + str(start_row) + ':' + 'K' + str(start_row)
        sheet.merge_cells(cell_range)
        sheet.cell(row=start_row, column=1).value = header_msg
        cell_range = 'A' + str(start_row + 1) + ':'+'B'+str(start_row + 1)
        sheet.merge_cells(cell_range)
        sheet.cell(row=start_row + 1, column=1).value = 'Scrip Details'
        max_column = sheet.max_column
        curr_column = 3
        while curr_column < max_column:
            sheet.merge_cells(start_row=start_row + 1, start_column=curr_column, end_row=start_row + 1,
                              end_column=curr_column + 1)
            curr_column = curr_column + 2

        sheet.merge_cells('C2:D2')
        sheet.cell(row=start_row + 1, column=3).value = 'Last Month'
        sheet.merge_cells('E2:F2')
        sheet.cell(row=start_row + 1, column=5).value = 'Current Month'
        sheet.merge_cells('G2:H2')
        sheet.cell(row=start_row + 1, column=7).value = 'Last Week'
        sheet.merge_cells('I2:J2')
        sheet.cell(row=start_row + 1, column=9).value = 'Current Week'

        curr_column = 1
        max_rows = sheet.max_row
        curr_row = start_row
        while curr_row <= max_rows:
            while curr_column <= max_column:
                sheet.cell(curr_row, curr_column).border = srgh.thin_border
                if curr_row == start_row + 1 or curr_row == start_row + 2:
                    sheet.cell(curr_row, curr_column).font = srgh.font_header
                    sheet.cell(curr_row, curr_column).fill = PatternFill(start_color='D3D3D3', fill_type="solid")
                    sheet.cell(curr_row, curr_column).alignment = srgh.align_header
                elif curr_row == start_row:
                    sheet.cell(curr_row, curr_column).font = srgh.font_header
                    if is_increased:
                        sheet.cell(curr_row, curr_column).fill = PatternFill(start_color='CAFF33', fill_type="solid")
                    else:
                        sheet.cell(curr_row, curr_column).fill = PatternFill(start_color='F6646B', fill_type="solid")
                    sheet.cell(curr_row, curr_column).alignment = srgh.align_header
                else:
                    sheet.cell(curr_row, curr_column).font = srgh.font_body
                    if curr_column < 3:
                        sheet.cell(curr_row, curr_column).alignment = srgh.align_body_str
                    else:
                        sheet.cell(curr_row, curr_column).alignment = srgh.align_body_num
                curr_column = curr_column + 1
            sheet.row_dimensions[curr_row].height = 20  # In pixels
            curr_row = curr_row + 1
            curr_column = 1

        curr_row_no = start_row + 3
        for rows in sheet.iter_rows(min_row=start_row + 3, max_row=max_rows, min_col=1):
            for cell in rows:
                if curr_row_no % 2 == 1:
                    cell.fill = PatternFill(start_color="f7ec8f", fill_type="solid")
                else:
                    cell.fill = PatternFill(start_color="edead7", fill_type="solid")
            curr_row_no = curr_row_no + 1
        sheet.column_dimensions['B'].width = 25
        sheet.column_dimensions['A'].width = 12

        book.save(report_name)

    # Format Cons Increase Report
    def format_close_above_last_week(self, report_name, header_msg):

        book = load_workbook(report_name)
        sheet = book['Trending Technical - 2']
        sheet.merge_cells('A1:K1')
        sheet.cell(row=1, column=1).value = header_msg
        sheet.merge_cells('A2:B2')
        sheet.cell(row=2, column=1).value = 'Scrip Details'
        sheet.merge_cells('C2:D2')
        sheet.cell(row=2, column=3).value = 'Last Month'
        sheet.merge_cells('E2:F2')
        sheet.cell(row=2, column=5).value = 'Current Month'
        sheet.merge_cells('G2:H2')
        sheet.cell(row=2, column=7).value = 'Last Week'
        sheet.merge_cells('I2:J2')
        sheet.cell(row=2, column=9).value = 'Current Week'

        max_column = sheet.max_column
        curr_column = 4

        curr_column = 1
        max_rows = sheet.max_row
        curr_row = 1
        while curr_row <= max_rows:
            while curr_column <= max_column:
                sheet.cell(curr_row, curr_column).border = srgh.thin_border
                if curr_row == 3 or curr_row == 2:
                    sheet.cell(curr_row, curr_column).font = srgh.font_header
                    sheet.cell(curr_row, curr_column).fill = PatternFill(start_color='D3D3D3', fill_type="solid")
                    sheet.cell(curr_row, curr_column).alignment = srgh.align_header
                elif curr_row == 1:
                    sheet.cell(curr_row, curr_column).font = srgh.font_header
                    sheet.cell(curr_row, curr_column).fill = PatternFill(start_color='CAFF33', fill_type="solid")
                    sheet.cell(curr_row, curr_column).alignment = srgh.align_header
                else:
                    sheet.cell(curr_row, curr_column).font = srgh.font_body
                    if curr_column < 3:
                        sheet.cell(curr_row, curr_column).alignment = srgh.align_body_str
                    else:
                        sheet.cell(curr_row, curr_column).alignment = srgh.align_body_num
                curr_column = curr_column + 1
            sheet.row_dimensions[curr_row].height = 20  # In pixels
            curr_row = curr_row + 1
            curr_column = 1

        curr_row_no = 4
        for rows in sheet.iter_rows(min_row=4, max_row=max_rows, min_col=1):
            for cell in rows:
                if curr_row_no % 2 == 1:
                    cell.fill = PatternFill(start_color="f7ec8f", fill_type="solid")
                else:
                    cell.fill = PatternFill(start_color="edead7", fill_type="solid")
            curr_row_no = curr_row_no + 1
        sheet.column_dimensions['A'].width = 18
        sheet.column_dimensions['B'].width = 25
        sheet.column_dimensions['C'].width = 12
        sheet.column_dimensions['D'].width = 12
        sheet.column_dimensions['E'].width = 12
        sheet.column_dimensions['F'].width = 12
        sheet.column_dimensions['G'].width = 12
        sheet.column_dimensions['H'].width = 12
        sheet.column_dimensions['I'].width = 12
        sheet.column_dimensions['J'].width = 12
        sheet.column_dimensions['K'].width = 15
        book.save(report_name)

    def generate_ca_records(self, report_name):

        # Read the Price File
        try:
            input_file_name = self.config.input_file_path + 'Bc' + self.current_date_str + '.csv'
            ca_details = pd.read_csv(input_file_name)
        except FileNotFoundError:
            print("Corporate Actions: file not found")
            return
        except EmptyDataError:
            print("There are no corporate actions today")
            return

        # Read the Scrip List
        scrip_list = pd.read_excel(self.config.master_scrip_list, "SCRIP_LIST")

        # Prepare the list for Selected List
        selected_list = pd.merge(scrip_list, ca_details, left_on=["SYMBOL", "SERIES"], right_on=["SYMBOL", "SERIES"])
        selected_list = selected_list[['SYMBOL', 'NAME', 'PURPOSE', 'RECORD_DT', 'EX_DT']]

        ca_details = ca_details[(ca_details['SERIES'] == 'EQ')]
        all_ca_list = ca_details[['SYMBOL', 'SECURITY', 'PURPOSE', 'RECORD_DT', 'EX_DT']]
        all_ca_list.rename(columns=srgh.ca_header_updated, inplace=True)

        narration1 = pd.DataFrame({'SYMBOL': 'Corporate Actions - My Scrips', 'NAME': ' ',
                               'PURPOSE': '', 'RECORD_DT': '', 'EX_DT': ''},
                              index=[0])

        narration2 = pd.DataFrame({'SYMBOL': 'Corporate Actions - Market', 'NAME': ' ',
                                   'PURPOSE': '', 'RECORD_DT': '', 'EX_DT': ''},
                                  index=[0])

        cust_header = pd.DataFrame({'SYMBOL': 'SYMBOL', 'NAME': 'Name', 'PURPOSE': 'Purpose',
                                'RECORD_DT': 'Record Date', 'EX_DT': 'Ex Date'}, index=[0])

        # Append Data Frame when Price Increased and Volume Increased
        narration1 = pd.concat([narration1, cust_header])
        selected_list = pd.concat([narration1, selected_list], sort=False)

        narration2 = pd.concat([narration2, cust_header])
        all_ca_list = pd.concat([narration2, all_ca_list], sort=False)
        selected_list = pd.concat([selected_list, all_ca_list], sort=False)

        book = load_workbook(report_name)
        writer = pd.ExcelWriter(report_name, engine='openpyxl')
        writer.book = book
        writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
        selected_list.to_excel(writer, "Corporate Actions", header=None, index=False)

        # Get the openpyxl workbook and worksheet objects.
        worksheet = writer.sheets['Corporate Actions']

        max_rows = worksheet.max_row
        st_row = 1
        st_col = 1
        for row_cells in worksheet.iter_rows(min_row=1, max_row=max_rows):
            for cell in row_cells:
                worksheet.cell(st_row, st_col).border = srgh.thin_border
                if 'Corporate Actions' in str(cell.value):
                    worksheet.merge_cells(start_row=st_row, start_column=st_col, end_row=st_row, end_column=st_col + 4)
                    worksheet.cell(st_row, st_col).fill = PatternFill(start_color='D3D3D3', fill_type="solid")
                    worksheet.cell(st_row, st_col).font = srgh.font_header
                    worksheet.cell(st_row, st_col).alignment = srgh.align_header
                    break
                elif str(cell.value) in srgh.ca_header_updated_1:
                    worksheet.cell(st_row, st_col).fill = PatternFill(start_color='D3D3D3', fill_type="solid")
                    worksheet.cell(st_row, st_col).font = srgh.font_header
                    worksheet.cell(st_row, st_col).alignment = srgh.align_header
                    st_col = st_col + 1

                else:
                    worksheet.cell(st_row, st_col).font = srgh.font_body
                    if st_row % 2 == 1:
                        cell.fill = PatternFill(start_color="f7ec8f", fill_type="solid")
                    else:
                        cell.fill = PatternFill(start_color="edead7", fill_type="solid")

                    if type(cell.value) == str:
                        worksheet.cell(st_row, st_col).alignment = srgh.align_body_str
                    else:
                        worksheet.cell(st_row, st_col).alignment = srgh.align_body_num
                    st_col = st_col + 1
            st_col = 1
            st_row = st_row + 1
            worksheet.row_dimensions[st_row].height = 20  # In pixels

            worksheet.column_dimensions['A'].width = 15
            worksheet.column_dimensions['B'].width = 30
            worksheet.column_dimensions['C'].width = 50
            worksheet.column_dimensions['D'].width = 20
            worksheet.column_dimensions['E'].width = 20
        writer.save()

    def generate_bulk_deal_records(self, report_name):

        # Read the Price File
        try:
            bulk_deals = pd.read_csv(self.config.bulk_file_name)
        except FileNotFoundError:
            print("Bulk Deals: file not found")
            return

        # Read the Scrip List
        scrip_list = pd.read_excel(self.config.master_scrip_list, "SCRIP_LIST")

        # Prepare the list for Selected List
        selected_list = pd.merge(scrip_list, bulk_deals, left_on=["SYMBOL"], right_on=["Symbol"])
        selected_list = selected_list[['Symbol', 'Security Name', 'Client Name', 'Buy/Sell', 'Quantity Traded',
                                       'Trade Price / Wght. Avg. Price']]

        all_bulk_deals = bulk_deals[['Symbol', 'Security Name', 'Client Name', 'Buy/Sell', 'Quantity Traded',
                                     'Trade Price / Wght. Avg. Price']]

        narration1 = pd.DataFrame({'Symbol': 'Bulk Deals - My Scrips', 'Security Name': ' ', 'Client Name': '',
                                   'Buy/Sell': '', 'Quantity Traded': '', 'Trade Price / Wght. Avg. Price': ''},
                                  index=[0])

        narration2 = pd.DataFrame({'Symbol': 'Bulk Deals - Market', 'Security Name': ' ', 'Client Name': '',
                                   'Buy/Sell': '', 'Quantity Traded': '', 'Trade Price / Wght. Avg. Price': ''},
                                  index=[0])

        cust_header = pd.DataFrame({'Symbol': 'Symbol', 'Security Name': 'Security Name', 'Client Name': 'Client Name',
                                    'Buy/Sell': 'Buy/Sell', 'Quantity Traded': 'Quantity Traded',
                                    'Trade Price / Wght. Avg. Price': 'Trade Price / Wght. Avg. Price'}, index=[0])

        # Append Data Frame when Price Increased and Volume Increased
        narration1 = pd.concat([narration1, cust_header])
        selected_list = pd.concat([narration1, selected_list], sort=False)

        narration2 = pd.concat([narration2, cust_header])
        all_bulk_deals = pd.concat([narration2, all_bulk_deals], sort=False)
        selected_list = pd.concat([selected_list, all_bulk_deals], sort=False)

        book = load_workbook(report_name)
        writer = pd.ExcelWriter(report_name, engine='openpyxl')
        writer.book = book
        writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
        selected_list.to_excel(writer, "Bulk Deals", header=None, index=False)

        # Get the openpyxl workbook and worksheet objects.
        worksheet = writer.sheets['Bulk Deals']

        max_rows = worksheet.max_row
        st_row = 1
        st_col = 1
        for row_cells in worksheet.iter_rows(min_row=1, max_row=max_rows):
            for cell in row_cells:
                worksheet.cell(st_row, st_col).border = srgh.thin_border
                if 'Bulk Deals' in str(cell.value):
                    worksheet.merge_cells(start_row=st_row, start_column=st_col, end_row=st_row, end_column=st_col + 5)
                    worksheet.cell(st_row, st_col).fill = PatternFill(start_color='D3D3D3', fill_type="solid")
                    worksheet.cell(st_row, st_col).font = srgh.font_header
                    worksheet.cell(st_row, st_col).alignment = srgh.align_header
                    break
                elif str(cell.value) in cust_header:
                    worksheet.cell(st_row, st_col).fill = PatternFill(start_color='D3D3D3', fill_type="solid")
                    worksheet.cell(st_row, st_col).font = srgh.font_header
                    worksheet.cell(st_row, st_col).alignment = srgh.align_header
                    st_col = st_col + 1

                else:
                    worksheet.cell(st_row, st_col).font = srgh.font_body
                    if st_row % 2 == 1:
                        cell.fill = PatternFill(start_color="f7ec8f", fill_type="solid")
                    else:
                        cell.fill = PatternFill(start_color="edead7", fill_type="solid")

                    if type(cell.value) == str:
                        worksheet.cell(st_row, st_col).alignment = srgh.align_body_str
                    else:
                        worksheet.cell(st_row, st_col).alignment = srgh.align_body_num
                    st_col = st_col + 1
            st_col = 1
            st_row = st_row + 1
            worksheet.row_dimensions[st_row].height = 20  # In pixels

            worksheet.column_dimensions['A'].width = 15
            worksheet.column_dimensions['B'].width = 35
            worksheet.column_dimensions['C'].width = 40
            worksheet.column_dimensions['D'].width = 10
            worksheet.column_dimensions['E'].width = 20
            worksheet.column_dimensions['F'].width = 30
        writer.save()

    def generate_block_deal_records(self, report_name):

        # Read the Price File
        try:
            block_deals = pd.read_csv(self.config.block_file_name)
        except FileNotFoundError:
            print("Block Deals: file not found")
            return

        block_deals = block_deals[['Symbol', 'Security Name', 'Client Name', 'Buy/Sell', 'Quantity Traded',
                                   'Trade Price / Wght. Avg. Price']]

        narration1 = pd.DataFrame({'Symbol': 'Block Deals', 'Security Name': ' ', 'Client Name': '',
                                   'Buy/Sell': '', 'Quantity Traded': '', 'Trade Price / Wght. Avg. Price': ''},
                                  index=[0])

        cust_header = pd.DataFrame({'Symbol': 'Symbol', 'Security Name': 'Security Name', 'Client Name': 'Client Name',
                                    'Buy/Sell': 'Buy/Sell', 'Quantity Traded': 'Quantity Traded',
                                    'Trade Price / Wght. Avg. Price': 'Trade Price / Wght. Avg. Price'}, index=[0])

        # Append Data Frame when Price Increased and Volume Increased
        narration1 = pd.concat([narration1, cust_header])
        selected_list = pd.concat([narration1, block_deals], sort=False)

        book = load_workbook(report_name)
        writer = pd.ExcelWriter(report_name, engine='openpyxl')
        writer.book = book
        writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
        selected_list.to_excel(writer, "Block Deals", header=None, index=False)

        # Get the openpyxl workbook and worksheet objects.
        worksheet = writer.sheets['Block Deals']

        max_rows = worksheet.max_row
        st_row = 1
        st_col = 1
        for row_cells in worksheet.iter_rows(min_row=1, max_row=max_rows):
            for cell in row_cells:
                worksheet.cell(st_row, st_col).border = srgh.thin_border
                if 'Block Deals' in str(cell.value):
                    worksheet.merge_cells(start_row=st_row, start_column=st_col, end_row=st_row, end_column=st_col + 5)
                    worksheet.cell(st_row, st_col).fill = PatternFill(start_color='D3D3D3', fill_type="solid")
                    worksheet.cell(st_row, st_col).font = srgh.font_header
                    worksheet.cell(st_row, st_col).alignment = srgh.align_header
                    break
                elif str(cell.value) in cust_header:
                    worksheet.cell(st_row, st_col).fill = PatternFill(start_color='D3D3D3', fill_type="solid")
                    worksheet.cell(st_row, st_col).font = srgh.font_header
                    worksheet.cell(st_row, st_col).alignment = srgh.align_header
                    st_col = st_col + 1

                else:
                    worksheet.cell(st_row, st_col).font = srgh.font_body
                    if st_row % 2 == 1:
                        cell.fill = PatternFill(start_color="f7ec8f", fill_type="solid")
                    else:
                        cell.fill = PatternFill(start_color="edead7", fill_type="solid")

                    if type(cell.value) == str:
                        worksheet.cell(st_row, st_col).alignment = srgh.align_body_str
                    else:
                        worksheet.cell(st_row, st_col).alignment = srgh.align_body_num
                    st_col = st_col + 1
            st_col = 1
            st_row = st_row + 1
            worksheet.row_dimensions[st_row].height = 20  # In pixels

            worksheet.column_dimensions['A'].width = 15
            worksheet.column_dimensions['B'].width = 35
            worksheet.column_dimensions['C'].width = 40
            worksheet.column_dimensions['D'].width = 10
            worksheet.column_dimensions['E'].width = 20
            worksheet.column_dimensions['F'].width = 30
        writer.save()
