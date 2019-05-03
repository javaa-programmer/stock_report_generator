import pandas as pd
from datetime import date
import stockreportgeneratorhelper as srgh
import directorypaths as dp
from openpyxl import load_workbook, formatting, styles
from openpyxl.styles.borders import BORDER_THIN
from openpyxl.styles import PatternFill, Border, Side
from openpyxl.utils import get_column_letter


class MasterReportUpdater:

    def __init__(self, input_file_name, sheet_name, current_date_str):
        self.input_file_name = input_file_name
        self.sheet_name = sheet_name
        self.current_date_str = current_date_str
        print(type(self.current_date_str))
        self.current_date = srgh.create_date(self.current_date_str)
        self.report_sheet_name = srgh.create_sheet_name(self.current_date_str)
        self.sheet_exist = srgh.check_sheet_exist(dp.master_report_name, self.report_sheet_name)

    # Reshape the Header
    # Creates the multi level header for master report
    def reshape_header(self, selected_list, excel_path, sheet_name):

        book = load_workbook(excel_path)
        writer = pd.ExcelWriter(excel_path, engine='openpyxl')
        writer.book = book
        writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
        # selected_list.to_excel(writer, sheet_name, header=None, index=False)

        if not self.sheet_exist:
            selected_list.loc[-1] = srgh.header3
        else:
            no_of_dyna_column = len(selected_list.columns) - len(srgh.header3)
            selected_list.loc[-1] = srgh.header3 + srgh.date_wise_dynamic_header * int(no_of_dyna_column / 6)

        selected_list.index = selected_list.index + 1  # shifting index
        selected_list = selected_list.sort_index()  # sorting by index

        selected_list.to_excel(writer, sheet_name, index=False)
        writer.save()

    # Get the openpyxl workbook and worksheet objects.
    # Format the excel rows and columns
    def format_final_excel(self):
        book = load_workbook(dp.master_report_name)
        sheet = book[self.report_sheet_name]
        sheet.merge_cells('A1:B1')
        sheet.merge_cells('C1:D1')
        sheet.merge_cells('E1:F1')
        sheet.merge_cells('G1:H1')
        sheet.merge_cells('I1:J1')
        sheet.merge_cells('K1:L1')

        max_column = sheet.max_column
        curr_column = 13
        while curr_column < max_column:
            sheet.merge_cells(start_row=1, start_column=curr_column, end_row=1, end_column=curr_column + 5)
            curr_column = curr_column + 6

        curr_column = 1
        max_rows = sheet.max_row
        curr_row = 1
        while curr_row <= max_rows:
            while curr_column <= max_column:
                sheet.cell(curr_row, curr_column).border = srgh.thin_border
                if curr_row == 1 or curr_row == 2:
                    sheet.cell(curr_row, curr_column).font = srgh.font_header
                    sheet.cell(curr_row, curr_column).fill = PatternFill(start_color='D3D3D3', fill_type="solid")
                    sheet.cell(curr_row, curr_column).alignment = srgh.align_header
                else:
                    sheet.cell(curr_row, curr_column).font = srgh.font_body
                    if curr_column < 3:
                        sheet.cell(curr_row, curr_column).alignment = srgh.align_body_str
                    else:
                        sheet.cell(curr_row, curr_column).alignment = srgh.align_body_num
                curr_column = curr_column + 1
            sheet.row_dimensions[curr_row].height = 20  # In pixels
            curr_row = curr_row + 1
            curr_column = 1

        curr_row_no = 3
        for rows in sheet.iter_rows(min_row=3, max_row=max_rows, min_col=1):
            for cell in rows:
                if curr_row_no % 2 == 1:
                    cell.fill = PatternFill(start_color="f7ec8f", fill_type="solid")
                else:
                    cell.fill = PatternFill(start_color="edead7", fill_type="solid")
            curr_row_no = curr_row_no + 1
        sheet.column_dimensions['B'].width = 25
        sheet.column_dimensions['A'].width = 12
        c = sheet['M4']
        sheet.freeze_panes = c

        start_row = 3
        end_row = max_rows
        column_counter = max_column

        while column_counter > 17:
            column_letter = get_column_letter((column_counter - 1))

            start_index = column_letter + str(start_row)
            end_index = column_letter + str(end_row)

            sheet.conditional_formatting.add(f'{start_index}:{end_index}',
                                             formatting.rule.CellIsRule(operator='lessThan',
                                                                        formula=['0'],
                                                                        fill=srgh.red_fill))

            sheet.conditional_formatting.add(f'{start_index}:{end_index}',
                                             formatting.rule.CellIsRule(operator='greaterThan', formula=['0'],
                                                                        fill=srgh.green_fill))
            column_counter = column_counter - 6

        book.save(dp.master_report_name)

    # update master report
    def update_master_report(self):
        mru = MasterReportUpdater(self.input_file_name, self.sheet_name, self.current_date_str)
        updated_record_set = mru.update_week_month_year()
        final_record_set = mru.calculate_month_weekly_high_low(updated_record_set,
                                                               srgh.create_date(self.current_date_str))
        date_wise_record_set = mru.update_date_wise_record(self.input_file_name, dp.master_report_name,
                                                           srgh.create_date(self.current_date_str), final_record_set,
                                                           self.current_date)
        mru.reshape_header(date_wise_record_set, dp.master_report_name, self.report_sheet_name)
        MasterReportUpdater.format_final_excel = staticmethod(MasterReportUpdater.format_final_excel)

        MasterReportUpdater.format_final_excel(self)

    # Update the month, year and week in the master data frame
    # For example, if date is 2019-04-15, following update will take place
    # Month = 4, Year = 2019 and week = 16
    # and returns the updated data frame
    def update_week_month_year(self):
        curr_week_data = pd.read_excel(self.input_file_name, self.sheet_name, parse_dates=['TRADE_DATE'])

        updated_month = pd.Series([])
        updated_year = pd.Series([])
        updated_week = pd.Series([])

        for index, row in curr_week_data.iterrows():
            trade_date = row['TRADE_DATE']
            updated_month[index] = row.TRADE_DATE.month
            year, week_number, weekday = date.replace(trade_date).isocalendar()
            updated_week[index] = week_number
            updated_year[index] = year

        curr_week_data.insert(4, 'MONTH', updated_month)
        curr_week_data.insert(5, 'WEEK', updated_week)
        curr_week_data.insert(6, 'YEAR', updated_year)

        return curr_week_data

    # Calculate the weekly High, Low and Monthly High and Low for new sheet.
    # If sheet already exist for any month, do nothing
    def calculate_month_weekly_high_low(self, df_scrip_list, current_date):
        sheet_exists = srgh.check_sheet_exist(dp.master_report_name, srgh.create_sheet_name(srgh.current_date_str))
        if sheet_exists:
            print("Sheet Exist...will not be created...")
            return

        print("Sheet does not exist... will create new")

        # Calculate Weekly High Price
        weekly_high_price = df_scrip_list.loc[df_scrip_list.groupby(['SYMBOL', 'WEEK', 'YEAR'])["HIGH_PRICE"].idxmax()]
        selected_fields_weekly_high = weekly_high_price[['SYMBOL', 'SERIES', 'WEEK', 'YEAR', 'HIGH_PRICE']]

        last_week_high_price = {}
        for index, row in selected_fields_weekly_high.iterrows():
            key = row.SYMBOL + str(row.WEEK + 1) + str(row.YEAR)
            last_week_high_price.update({key: row.HIGH_PRICE})

        for index, row in selected_fields_weekly_high.iterrows():
            key = row.SYMBOL + str(row.WEEK) + str(row.YEAR)
            selected_fields_weekly_high.loc[index, 'LAST_WEEK_HIGH_PRICE'] = last_week_high_price.get(key)

        updated_record_set_weekly_high = pd.merge(df_scrip_list, selected_fields_weekly_high,
                                              on=['SYMBOL', 'SERIES', 'WEEK', 'YEAR'])

        # Calculate Weekly High Price
        weekly_low_price = df_scrip_list.loc[df_scrip_list.groupby(['SYMBOL', 'WEEK', 'YEAR'])["LOW_PRICE"].idxmin()]
        selected_fields_weekly_low = weekly_low_price[['SYMBOL', 'SERIES', 'WEEK', 'YEAR', 'LOW_PRICE']]

        # last_week_low_price
        last_week_low_price = {}
        for index, row in selected_fields_weekly_low.iterrows():
            key = row.SYMBOL + str(row.WEEK + 1) + str(row.YEAR)
            last_week_low_price.update({key: row.LOW_PRICE})

        for index, row in selected_fields_weekly_low.iterrows():
            key = row.SYMBOL + str(row.WEEK) + str(row.YEAR)
            selected_fields_weekly_low.loc[index, 'LAST_WEEK_LOW_PRICE'] = last_week_low_price.get(key)

        updated_record_set_weekly_high_low = pd.merge(updated_record_set_weekly_high, selected_fields_weekly_low,
                                                  on=['SYMBOL', 'SERIES', 'WEEK', 'YEAR'])

        # Calculate Monthly High Price
        monthly_high_price = df_scrip_list.loc[df_scrip_list.groupby(['SYMBOL', 'MONTH', 'YEAR'])["HIGH_PRICE"].idxmax()]
        selected_fields_monthly_high = monthly_high_price[['SYMBOL', 'SERIES', 'MONTH', 'YEAR', 'HIGH_PRICE']]

        # last_month_high_price
        last_month_high_price = {}
        for index, row in selected_fields_monthly_high.iterrows():
            key = row.SYMBOL + str(row.MONTH + 1) + str(row.YEAR)
            last_month_high_price.update({key: row.HIGH_PRICE})

        for index, row in selected_fields_monthly_high.iterrows():
            key = row.SYMBOL + str(row.MONTH) + str(row.YEAR)
            selected_fields_monthly_high.loc[index, 'LAST_MO_HIGH_PRICE'] = last_month_high_price.get(key)

        updated_record_set_monthly_high = pd.merge(updated_record_set_weekly_high_low, selected_fields_monthly_high,
                                               on=['SYMBOL', 'SERIES', 'MONTH', 'YEAR'])

        # Calculate Monthly Low Price
        monthly_high_price = df_scrip_list.loc[df_scrip_list.groupby(['SYMBOL', 'MONTH', 'YEAR'])["LOW_PRICE"].idxmin()]
        selected_fields_monthly_high = monthly_high_price[['SYMBOL', 'SERIES', 'MONTH', 'YEAR', 'LOW_PRICE']]

        # last_month_low_price
        last_month_low_price = {}
        for index, row in selected_fields_monthly_high.iterrows():
            key = row.SYMBOL + str(row.MONTH + 1) + str(row.YEAR)
            last_month_low_price.update({key: row.LOW_PRICE})

        for index, row in selected_fields_monthly_high.iterrows():
            key = row.SYMBOL + str(row.MONTH) + str(row.YEAR)
            selected_fields_monthly_high.loc[index, 'LAST_MO_LOW_PRICE'] = last_month_low_price.get(key)

        updated_record_set_monthly_high_low = pd.merge(updated_record_set_monthly_high, selected_fields_monthly_high,
                                                   on=['SYMBOL', 'SERIES', 'MONTH', 'YEAR'])

        updated_record_set_monthly_high_low.rename(columns={'HIGH_PRICE': 'MO_HIGH_PRICE', 'HIGH_PRICE_x': 'HIGH_PRICE',
                                                        'HIGH_PRICE_y': 'WE_HIGH_PRICE'}, inplace=True)

        updated_record_set_monthly_high_low.rename(columns={'LOW_PRICE': 'MO_LOW_PRICE', 'LOW_PRICE_x': 'LOW_PRICE',
                                                        'LOW_PRICE_y': 'WE_LOW_PRICE'}, inplace=True)

        final_updated_records = updated_record_set_monthly_high_low[updated_record_set_monthly_high_low['TRADE_DATE']
                                                                == current_date]

        fl_record_set = final_updated_records[['SYMBOL', 'NAME', 'HI_52_WK', 'LO_52_WK', 'LAST_MO_HIGH_PRICE',
                                           'LAST_MO_LOW_PRICE', 'MO_HIGH_PRICE', 'MO_LOW_PRICE',
                                           'LAST_WEEK_HIGH_PRICE', 'LAST_WEEK_LOW_PRICE', 'WE_HIGH_PRICE',
                                           'WE_LOW_PRICE']]

        return fl_record_set

    # Rename the newly generated columns once the columns are added for run date
    def rename_date_wise_column(self, existing_header_values):
        column_date_dict = {}
        counter = 0

        for val in existing_header_values:
            if counter == 0:
                column_date_dict.update({'Open': val.date()})
            else:
                column_date_dict.update({'Open.'+str(counter): val.date()})
            counter = counter + 1
        return column_date_dict

    # Updates the Open, High, Low and Close price for Current Date
    # Same time update the Weekly High Low and Monthly High Low price as well
    # If the date is first day of week, calculate last week High and Low price as well
    def update_date_wise_record(self, master_data, master_report_name, sheet_name, master_report_data, current_date):
        MasterReportUpdater.rename_date_wise_column = staticmethod(MasterReportUpdater.rename_date_wise_column)
        curr_week_data = pd.read_excel(self.input_file_name, 'Details')
        curr_week_data = curr_week_data[curr_week_data['TRADE_DATE'] == current_date]
        curr_week_data = curr_week_data[['SYMBOL', 'HI_52_WK', 'LO_52_WK', 'PREV_CL_PR', 'OPEN_PRICE', 'HIGH_PRICE',
                                     'LOW_PRICE', 'CLOSE_PRICE', 'NET_TRDQTY']]
        existing_header_values = []

        is_new_week = srgh.check_new_week(self.current_date)

        if self.sheet_exist:
            monthly_report = pd.read_excel(dp.master_report_name, self.report_sheet_name, skiprows=1)
        else:
            monthly_report = master_report_data

        # Prepare the list for Selected List
        selected_list = pd.merge(monthly_report, curr_week_data, left_on=["SYMBOL"], right_on=["SYMBOL"], how='right')

        no_cols = len(selected_list.columns)

        change_per = (((selected_list[selected_list.columns[no_cols - 2]] - selected_list[
            selected_list.columns[no_cols - 6]]) * 100)
                  / selected_list[selected_list.columns[no_cols - 6]])

        selected_list.insert((len(selected_list.columns) - 1), "Change", change_per)

        if self.sheet_exist:

            df = pd.read_excel(dp.master_report_name, self.report_sheet_name)
            header_list = list(df.columns.values)
            counter = 12

            while counter < len(header_list):
                existing_header_values.append(header_list[counter])
                counter = counter + 6

            for i, row in selected_list.iterrows():
                if row['High.2'] < row['HIGH_PRICE']:
                    selected_list.at[i, 'High.2'] = row['HIGH_PRICE']

                if row['Low.2'] > row['LOW_PRICE']:
                    selected_list.at[i, 'Low.2'] = row['LOW_PRICE']

                if not is_new_week:
                    if row['High.4'] < row['HIGH_PRICE']:
                        selected_list.at[i, 'High.4'] = row['HIGH_PRICE']

                    if row['Low.4'] > row['LOW_PRICE']:
                        selected_list.at[i, 'Low.4'] = row['LOW_PRICE']
                else:
                    selected_list.at[i, 'High.3'] = row['High.4']
                    selected_list.at[i, 'Low.3'] = row['Low.4']
                    selected_list.at[i, 'High.4'] = row['HIGH_PRICE']
                    selected_list.at[i, 'Low.4'] = row['LOW_PRICE']

            selected_list[selected_list.columns[2]] = selected_list.HI_52_WK
            selected_list[selected_list.columns[3]] = selected_list.LO_52_WK

            selected_list.drop(columns=['HI_52_WK', 'LO_52_WK', 'PREV_CL_PR'], inplace=True)
            selected_list.rename(columns=srgh.existing_header, inplace=True)
            selected_list.rename(columns=MasterReportUpdater.rename_date_wise_column(self, existing_header_values), inplace=True)
            selected_list.rename(columns=srgh.dynamic_header, inplace=True)
            selected_list.rename(columns={'OPEN_PRICE': current_date.date()}, inplace=True)
        else:
            selected_list.HI_52_WK_x.replace(selected_list.HI_52_WK_y, inplace=True)
            selected_list.LO_52_WK_x.replace(selected_list.LO_52_WK_x, inplace=True)
            selected_list.drop(columns=['HI_52_WK_y', 'LO_52_WK_y', 'PREV_CL_PR'], inplace=True)
            selected_list.rename(columns=srgh.header1, inplace=True)
            selected_list.rename(columns=srgh.dynamic_header, inplace=True)
            selected_list.rename(columns={'OPEN_PRICE': current_date.date()}, inplace=True)

        # Create a new variable called 'header' from the first row of the dataset
        return selected_list
